from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.utils.column_mapping_utils import normalize_header
from app.utils.excel_raw_value import format_stored_raw_value, serialize_raw_excel_value

TARGET_NUMERIC_COLUMNS = {
    "dn_mm",
    "id_mm",
    "od_mm",
    "wall_thk_mm",
    "weight",
    "half_od_mm",
    "area_m2_per_m",
    "sch_mm",
    "dm_ex",
    "radius",
}


def _header_column_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        raw_header = str(cell.value).strip()
        if not raw_header:
            continue
        normalized = normalize_header(raw_header)
        mapping[normalized] = col_idx
    return mapping


def debug_excel_raw_cells(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 10,
) -> list[dict[str, Any]]:
    """
    Lê valores crus do Excel via openpyxl, sem parser decimal.
    Retorna metadados de célula (tipo, data_type, number_format).
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {file_path}")

    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        header_to_col = _header_column_map(ws)
        target_headers = [h for h in header_to_col if h in TARGET_NUMERIC_COLUMNS]

        result: list[dict[str, Any]] = []
        last_row = min(ws.max_row or 1, max_rows + 1)

        for row_idx in range(2, last_row + 1):
            row_info: dict[str, Any] = {
                "excel_row_number": row_idx,
                "cells": [],
            }

            for header in target_headers:
                col_idx = header_to_col[header]
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value

                row_info["cells"].append(
                    {
                        "excel_column": header,
                        "cell_coordinate": cell.coordinate,
                        "value": value,
                        "value_str": None if value is None else str(value),
                        "python_type": type(value).__name__ if value is not None else None,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "serialized_for_staging": serialize_raw_excel_value(value),
                    }
                )

            if row_info["cells"]:
                result.append(row_info)

        return result
    finally:
        wb.close()


def _values_equivalent(excel_value: Any, staging_value: Any) -> bool:
    excel_text = format_stored_raw_value(serialize_raw_excel_value(excel_value))
    staging_text = format_stored_raw_value(staging_value)
    if excel_text is None and staging_text is None:
        return True
    if excel_text is None or staging_text is None:
        return False
    return excel_text == staging_text


def _is_likely_contaminated(value: Any, column_name: str) -> bool:
    from app.services.staging_integrity import is_likely_corrupted_raw_numeric

    return is_likely_corrupted_raw_numeric(value, column_name)


def compare_sample_status(
    *,
    column: str,
    excel_value: Any,
    staging_value: Any,
    excel_present: bool,
    staging_present: bool,
) -> str:
    if not excel_present and staging_present:
        return "MISSING_IN_EXCEL"
    if excel_present and not staging_present:
        return "MISSING_IN_STAGING"
    if not excel_present and not staging_present:
        return "MISSING_IN_EXCEL"

    if _values_equivalent(excel_value, staging_value):
        if _is_likely_contaminated(excel_value, column) or _is_likely_contaminated(staging_value, column):
            return "OK_CONTAMINATED_SOURCE"
        return "OK"

    return "MISMATCH"


def infer_diagnosis(samples: list[dict[str, Any]]) -> str:
    statuses = {sample["status"] for sample in samples}
    if not samples:
        return "NO_SAMPLES"
    if statuses <= {"OK"}:
        return "OK"
    if "MISMATCH" in statuses:
        return "STAGING_BUG"
    if statuses <= {"OK_CONTAMINATED_SOURCE", "OK"}:
        return "CONTAMINATED_SOURCE"
    if "MISMATCH" in statuses and "OK_CONTAMINATED_SOURCE" in statuses:
        return "MIXED"
    return "MIXED"
