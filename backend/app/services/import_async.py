import json
import re
import time
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import text

from app.core.database import SessionLocal
from app.core.logging import log_error, log_event
from app.services import import_staging_service as staging
from app.services import spec_items_service
from app.services.import_flow_service import (
    _build_mapped_row,
    _column_metadata_map,
    _column_nullable_map,
    _generate_preview,
)
from fastapi import HTTPException

from app.services.import_guard import validate_preview_ready_for_apply
from app.services.import_schema import ImportSchemaError, require_import_schema
from app.services.import_staging_reader import build_raw_row_json, is_empty_excel_row, normalize_excel_headers
from app.services.import_run_lock import release_run_lock, try_acquire_run_lock
from app.services.job_runner import submit_import_job
from app.utils.column_mapping_utils import suggest_column_mapping
from app.utils.excel_utils import detect_sheet_name, serialize_for_compare, utc_now
from app.utils.value_parser import NUMERIC_COLUMNS, validate_numeric_range

STORAGE_DIR = Path(__file__).resolve().parents[1] / "storage" / "imports"
PROGRESS_LOG_EVERY = 1000
STAGING_BATCH = 500


def _ensure_staging_not_corrupted(db, run_id: int) -> None:
    """Legado: não bloqueia mais por valores numéricos grandes — coerção inteligente trata no preview."""
    return


def _build_mapping_records(raw_headers: list[str], db_columns: list[str]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    db_set = set(db_columns)
    mapping_records: list[dict[str, Any]] = []
    known_columns: list[str] = []
    unknown_columns: list[str] = []

    for excel_col in raw_headers:
        if not excel_col:
            continue
        suggestion = suggest_column_mapping(excel_col, db_columns)
        if excel_col in db_set:
            known_columns.append(excel_col)
            mapping_records.append(
                {
                    "excel_column_name": excel_col,
                    "target_column_name": excel_col,
                    "action": "MAP_TO_EXISTING",
                    "confidence": 1.0,
                }
            )
        else:
            unknown_columns.append(excel_col)
            target = suggestion.get("suggested_target_column")
            action = suggestion.get("action", "IGNORE")
            if target and action == "MAP_TO_EXISTING":
                mapping_records.append(
                    {
                        "excel_column_name": excel_col,
                        "target_column_name": target,
                        "action": "MAP_TO_EXISTING",
                        "confidence": suggestion.get("confidence", 0.95),
                    }
                )
            else:
                mapping_records.append(
                    {
                        "excel_column_name": excel_col,
                        "target_column_name": None,
                        "action": "IGNORE",
                        "confidence": 0.0,
                    }
                )

    return mapping_records, known_columns, unknown_columns


def _stage_workbook_rows(
    db,
    run_id: int,
    file_path: str,
    *,
    estimated_total_hint: int | None = None,
) -> tuple[str, int, list[str], list[str]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    log_event("EXCEL_IMPORT", "workbook_loaded", run_id=run_id, sheets=sheet_names)

    sheet_name = detect_sheet_name(sheet_names)
    if not sheet_name:
        wb.close()
        raise ValueError(f"Não foi possível detectar aba. Abas: {', '.join(sheet_names)}")

    log_event("EXCEL_IMPORT", "sheet_selected", run_id=run_id, sheet_name=sheet_name)
    ws = wb[sheet_name]
    estimated_total = estimated_total_hint or max((ws.max_row or 1) - 1, 0)

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("Planilha vazia.")

    raw_headers = normalize_excel_headers(header_row)
    log_event("EXCEL_IMPORT", "headers_detected", run_id=run_id, headers=raw_headers)

    db_columns = spec_items_service.get_valid_column_names(db)
    mapping_records, known_columns, unknown_columns = _build_mapping_records(raw_headers, db_columns)

    log_event(
        "EXCEL_IMPORT",
        "mapping_suggested",
        run_id=run_id,
        unknown_columns=unknown_columns,
    )

    staging.update_import_progress(
        db,
        run_id,
        status="ANALYZING",
        phase="raw_staging",
        message="Salvando linhas em staging (valores crus)...",
    )
    log_event("EXCEL_IMPORT", "raw_staging_started", run_id=run_id)

    batch: list[dict[str, Any]] = []
    rows_read = 0
    staged_rows = 0

    for row_idx, row in enumerate(rows_iter, start=2):
        if is_empty_excel_row(row):
            continue

        row_data = build_raw_row_json(raw_headers, row)
        batch.append({"excel_row_number": row_idx, "raw_json": row_data})
        rows_read += 1

        if len(batch) >= STAGING_BATCH:
            staged_rows += staging.save_staging_rows(db, run_id, batch, commit_batch_size=STAGING_BATCH)
            batch = []
            if rows_read % PROGRESS_LOG_EVERY == 0:
                log_event("EXCEL_IMPORT", "raw_staging_progress", run_id=run_id, staged_rows=rows_read)
                staging.update_import_progress(
                    db,
                    run_id,
                    status="ANALYZING",
                    phase="raw_staging",
                    current=rows_read,
                    total=max(estimated_total, rows_read),
                    message=f"Salvando staging cru... ({rows_read}/{max(estimated_total, rows_read)})",
                    commit=True,
                )

    if batch:
        staged_rows += staging.save_staging_rows(db, run_id, batch, commit_batch_size=STAGING_BATCH)

    wb.close()
    staging.save_column_mappings(db, run_id, mapping_records)

    db.execute(
        text(
            """
            UPDATE app_import_runs
            SET sheet_name = :sheet_name, total_rows = :total_rows, finished_at = now()
            WHERE id = :id
            """
        ),
        {"sheet_name": sheet_name, "total_rows": staged_rows, "id": run_id},
    )
    db.commit()

    return sheet_name, staged_rows, known_columns, unknown_columns


def _job_db():
    return SessionLocal()


def _safe_filename(name: str) -> str:
    return re.sub(r"[^\w.\-]", "_", name) or "upload.xlsx"


def _save_upload_file(upload_file: UploadFile, run_id: int) -> Path:
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    file_name = upload_file.filename or "upload.xlsx"
    file_path = STORAGE_DIR / f"{run_id}_{_safe_filename(file_name)}"
    return file_path


async def _stream_upload_to_disk(upload_file: UploadFile, file_path: Path) -> None:
    with file_path.open("wb") as output:
        while True:
            chunk = await upload_file.read(1024 * 1024)
            if not chunk:
                break
            output.write(chunk)


async def start_analyze_from_upload(db, upload_file: UploadFile) -> dict[str, Any]:
    require_import_schema(db)

    request_started = time.perf_counter()
    file_name = upload_file.filename or "upload.xlsx"
    log_event(
        "EXCEL_IMPORT",
        "analyze_request_received",
        file_name=file_name,
    )

    run_id = staging.create_import_run(
        db,
        file_name=file_name,
        status="ANALYZING",
        source_file_path=None,
    )
    file_path = _save_upload_file(upload_file, run_id)
    await _stream_upload_to_disk(upload_file, file_path)

    db.execute(
        text("UPDATE app_import_runs SET source_file_path = :path WHERE id = :id"),
        {"path": str(file_path), "id": run_id},
    )
    staging.update_import_progress(
        db,
        run_id,
        status="ANALYZING",
        phase="queued",
        current=0,
        total=0,
        message="Arquivo recebido. Análise enfileirada...",
        commit=False,
    )
    db.commit()

    submit_import_job("excel_analyze", run_id, run_analyze_job, run_id, str(file_path))

    duration_ms = round((time.perf_counter() - request_started) * 1000, 2)
    log_event(
        "EXCEL_IMPORT",
        "analyze_request_completed",
        run_id=run_id,
        saved_file_path=str(file_path),
        duration_ms=duration_ms,
    )

    return {
        "ok": True,
        "run_id": run_id,
        "status": "ANALYZING",
        "message": "Análise iniciada. Acompanhe pelo status.",
    }


def start_analyze(db, file_bytes: bytes, file_name: str) -> dict[str, Any]:
    """Compatibilidade para testes/scripts: grava bytes e enfileira job."""
    require_import_schema(db)

    request_started = time.perf_counter()
    log_event(
        "EXCEL_IMPORT",
        "analyze_request_received",
        file_name=file_name,
    )

    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    run_id = staging.create_import_run(
        db,
        file_name=file_name,
        status="ANALYZING",
        source_file_path=None,
    )
    file_path = STORAGE_DIR / f"{run_id}_{_safe_filename(file_name)}"
    file_path.write_bytes(file_bytes)

    db.execute(
        text("UPDATE app_import_runs SET source_file_path = :path WHERE id = :id"),
        {"path": str(file_path), "id": run_id},
    )
    staging.update_import_progress(
        db,
        run_id,
        status="ANALYZING",
        phase="queued",
        current=0,
        total=0,
        message="Arquivo recebido. Análise enfileirada...",
        commit=False,
    )
    db.commit()

    submit_import_job("excel_analyze", run_id, run_analyze_job, run_id, str(file_path))

    duration_ms = round((time.perf_counter() - request_started) * 1000, 2)
    log_event(
        "EXCEL_IMPORT",
        "analyze_request_completed",
        run_id=run_id,
        saved_file_path=str(file_path),
        duration_ms=duration_ms,
    )

    return {
        "ok": True,
        "run_id": run_id,
        "status": "ANALYZING",
        "message": "Análise iniciada. Acompanhe pelo status.",
    }


def run_analyze_job(run_id: int, file_path: str | None = None) -> None:
    db = _job_db()
    started = time.time()
    try:
        run = staging.get_import_run(db, run_id)
        if not run:
            raise ValueError(f"Run {run_id} não encontrado")

        resolved_path = file_path or run.get("source_file_path")
        if not resolved_path or not Path(resolved_path).exists():
            raise ValueError("Arquivo de importação não encontrado")

        log_event("EXCEL_IMPORT", "analyze_started", run_id=run_id, file_name=run["file_name"])

        staging.update_import_progress(
            db,
            run_id,
            status="ANALYZING",
            phase="loading_workbook",
            message="Carregando planilha...",
        )

        sheet_name, staged_rows, known_columns, unknown_columns = _stage_workbook_rows(
            db,
            run_id,
            resolved_path,
        )

        final_status = "MAPPING_REQUIRED"
        staging.update_import_progress(
            db,
            run_id,
            status=final_status,
            phase="completed",
            current=staged_rows,
            total=staged_rows,
            message="Análise concluída. Revise o mapeamento de colunas.",
        )

        duration_ms = int((time.time() - started) * 1000)
        log_event(
            "EXCEL_IMPORT",
            "analyze_completed",
            run_id=run_id,
            total_rows=staged_rows,
            known_columns_count=len(known_columns),
            unknown_columns_count=len(unknown_columns),
            duration_ms=duration_ms,
        )
    except Exception as exc:
        db.rollback()
        staging.update_import_progress(
            db,
            run_id,
            status="FAILED",
            phase="failed",
            message=str(exc),
        )
        staging.update_run_status(db, run_id, "FAILED", error_message=str(exc))
        db.commit()
        log_error(
            "EXCEL_IMPORT",
            "analyze_failed",
            run_id=run_id,
            error=str(exc),
            error_type=type(exc).__name__,
        )
        raise
    finally:
        db.close()


def _normalize_mappings(mappings: list[dict[str, Any]], db_columns: set[str]) -> list[dict[str, Any]]:
    used_targets: set[str] = set()
    normalized: list[dict[str, Any]] = []
    for item in mappings:
        action = item.get("action", "IGNORE")
        excel_col = item["excel_column_name"]
        target = item.get("target_column_name")
        if action == "MAP_TO_EXISTING":
            if not target or target not in db_columns:
                raise ValueError(f"Coluna destino inválida para '{excel_col}'")
            if target in used_targets:
                raise ValueError(f"Coluna destino duplicada: {target}")
            used_targets.add(target)
        normalized.append(
            {
                "excel_column_name": excel_col,
                "target_column_name": target if action == "MAP_TO_EXISTING" else None,
                "action": action,
                "confidence": item.get("confidence"),
            }
        )
    return normalized


def save_mapping_only(db, run_id: int, mappings: list[dict[str, Any]]) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")
    if run["status"] == "APPLIED":
        raise ValueError("Importação já aplicada.")
    if run["status"] in {"PREVIEWING", "APPLYING", "ANALYZING"}:
        raise ValueError(f"Processamento em andamento (status={run['status']})")

    db_columns = set(spec_items_service.get_valid_column_names(db))
    normalized = _normalize_mappings(mappings, db_columns)
    staging.save_column_mappings(db, run_id, normalized)
    staging.update_import_progress(
        db,
        run_id,
        status="MAPPING_REQUIRED",
        phase="mapping_saved",
        message="Mapeamento salvo. Clique em Gerar preview.",
    )
    db.commit()
    log_event("EXCEL_IMPORT", "mapping_saved", run_id=run_id, mappings_count=len(normalized))
    return {
        "ok": True,
        "run_id": run_id,
        "status": "MAPPING_REQUIRED",
        "message": "Mapeamento salvo.",
    }


def start_preview(db, run_id: int, mappings: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    require_import_schema(db)

    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")

    current_status = run["status"]
    log_event(
        "EXCEL_IMPORT",
        "preview_request_received",
        run_id=run_id,
        current_status=current_status,
    )

    if current_status == "PREVIEWING":
        log_event(
            "EXCEL_IMPORT",
            "preview_request_ignored_already_running",
            run_id=run_id,
            current_status=current_status,
        )
        return {
            "ok": True,
            "run_id": run_id,
            "status": "PREVIEWING",
            "already_running": True,
            "message": "Preview já está em processamento.",
        }

    if current_status == "PREVIEW_READY":
        return {
            "ok": True,
            "run_id": run_id,
            "status": "PREVIEW_READY",
            "already_ready": True,
            "message": "Preview já está pronto.",
        }

    if current_status in {"APPLYING", "APPLIED"}:
        log_error(
            "EXCEL_IMPORT",
            "preview_request_rejected_invalid_status",
            run_id=run_id,
            current_status=current_status,
        )
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Não é possível gerar preview neste status.",
                "current_status": current_status,
            },
        )

    if current_status == "ANALYZING":
        log_event(
            "EXCEL_IMPORT",
            "preview_request_ignored_already_running",
            run_id=run_id,
            current_status=current_status,
        )
        return {
            "ok": True,
            "run_id": run_id,
            "status": current_status,
            "already_running": True,
            "message": "Análise ainda em andamento.",
        }

    if not try_acquire_run_lock(run_id):
        log_event(
            "EXCEL_IMPORT",
            "preview_request_ignored_already_running",
            run_id=run_id,
            current_status=current_status,
            reason="run_lock_busy",
        )
        return {
            "ok": True,
            "run_id": run_id,
            "status": current_status,
            "already_running": True,
            "message": "Já existe processamento em andamento para esta importação.",
        }

    try:
        if mappings is not None:
            db_columns = set(spec_items_service.get_valid_column_names(db))
            normalized = _normalize_mappings(mappings, db_columns)
            staging.save_column_mappings(db, run_id, normalized)

        locked = db.execute(
            text(
                """
                UPDATE app_import_runs
                SET status = 'PREVIEWING',
                    phase = 'queued',
                    progress_message = 'Preview enfileirado...'
                WHERE id = :run_id
                  AND status NOT IN ('PREVIEWING', 'APPLYING', 'APPLIED', 'ANALYZING')
                RETURNING id
                """
            ),
            {"run_id": run_id},
        ).scalar()

        if not locked:
            db.rollback()
            run = staging.get_import_run(db, run_id) or {}
            refreshed_status = run.get("status", current_status)
            if refreshed_status == "PREVIEWING":
                return {
                    "ok": True,
                    "run_id": run_id,
                    "status": "PREVIEWING",
                    "already_running": True,
                    "message": "Preview já está em processamento.",
                }
            if refreshed_status == "PREVIEW_READY":
                return {
                    "ok": True,
                    "run_id": run_id,
                    "status": "PREVIEW_READY",
                    "already_ready": True,
                    "message": "Preview já está pronto.",
                }
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Não foi possível iniciar preview neste status.",
                    "current_status": refreshed_status,
                },
            )

        staging.update_import_progress(
            db,
            run_id,
            status="PREVIEWING",
            phase="queued",
            current=0,
            message="Preview enfileirado...",
        )
        db.commit()

        log_event("EXCEL_IMPORT", "preview_job_starting", run_id=run_id)
        submit_import_job("excel_preview", run_id, run_preview_job, run_id)
        return {"ok": True, "run_id": run_id, "status": "PREVIEWING"}
    except Exception:
        release_run_lock(run_id)
        raise


def run_preview_job(run_id: int) -> None:
    db = _job_db()
    started = time.time()
    try:
        require_import_schema(db)
        log_event("EXCEL_IMPORT", "preview_started", run_id=run_id)
        total_rows = staging.count_staging_rows(db, run_id)
        staging.update_import_progress(
            db,
            run_id,
            status="PREVIEWING",
            phase="starting",
            current=0,
            total=total_rows,
            message="Preparando geração do preview...",
        )
        preview = _generate_preview(
            db,
            run_id,
            progress_callback=lambda c, t, m: staging.update_import_progress(
                db,
                run_id,
                status="PREVIEWING",
                phase="processing",
                current=c,
                total=t,
                message=m,
                commit=True,
            ),
        )
        if not preview.get("ok", True):
            status = preview.get("status", "FAILED")
            staging.update_run_status(
                db,
                run_id,
                status,
                error_message=preview.get("message"),
            )
            staging.update_import_progress(
                db,
                run_id,
                status=status,
                phase="failed",
                message=preview.get("message"),
            )
            db.commit()
            return

        staging.update_run_counts(
            db,
            run_id,
            inserted_rows=preview["summary"]["insert_rows"],
            updated_rows=preview["summary"]["update_rows"],
            unchanged_rows=preview["summary"]["unchanged_rows"],
            ignored_rows=preview["summary"]["ignored_rows"],
            error_rows=preview["summary"]["error_rows"],
            status="PREVIEW_READY",
        )
        staging.save_preview_meta(db, run_id, preview.get("preview_meta", {}))
        staging.update_import_progress(
            db,
            run_id,
            status="PREVIEW_READY",
            phase="completed",
            current=preview["summary"]["total_rows"],
            total=preview["summary"]["total_rows"],
            message="Preview pronto para revisão.",
        )
        db.commit()
        duration_ms = int((time.time() - started) * 1000)
        log_event(
            "EXCEL_IMPORT",
            "preview_completed",
            run_id=run_id,
            summary=preview["summary"],
            duration_ms=duration_ms,
        )
    except ImportSchemaError as exc:
        db.rollback()
        message = exc.detail.get("message", str(exc))
        staging.update_import_progress(db, run_id, status="FAILED", phase="failed", message=message)
        staging.update_run_status(db, run_id, "FAILED", error_message=message)
        db.commit()
        log_error(
            "EXCEL_IMPORT",
            "preview_schema_incomplete",
            run_id=run_id,
            missing_columns=exc.detail.get("missing_columns"),
            error=message,
        )
    except Exception as exc:
        db.rollback()
        staging.update_import_progress(db, run_id, status="FAILED", phase="failed", message=str(exc))
        staging.update_run_status(db, run_id, "FAILED", error_message=str(exc))
        db.commit()
        log_error("EXCEL_IMPORT", "preview_failed", run_id=run_id, error=str(exc), error_type=type(exc).__name__)
        raise
    finally:
        release_run_lock(run_id)
        db.close()


def start_reanalyze(db, run_id: int) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")
    if run["status"] == "APPLIED":
        raise ValueError("Importação já aplicada.")
    if run["status"] in {"ANALYZING", "PREVIEWING", "APPLYING"}:
        raise ValueError(f"Processamento em andamento (status={run['status']})")

    file_path = run.get("source_file_path")
    if not file_path or not Path(file_path).exists():
        return {
            "ok": False,
            "message": "Arquivo original não encontrado. Reenvie o Excel.",
        }

    staging.clear_preview_artifacts(db, run_id)
    staging.clear_staging_rows(db, run_id)
    staging.update_import_progress(
        db,
        run_id,
        status="ANALYZING",
        phase="queued",
        current=0,
        message="Reanálise enfileirada...",
    )
    db.commit()
    submit_import_job("excel_reanalyze", run_id, run_reanalyze_job, run_id)
    log_event("EXCEL_IMPORT", "reanalyze_queued", run_id=run_id)
    return {
        "ok": True,
        "run_id": run_id,
        "status": "ANALYZING",
        "message": "Reanalisando Excel original com staging cru.",
    }


def run_reanalyze_job(run_id: int) -> None:
    db = _job_db()
    started = time.time()
    try:
        run = staging.get_import_run(db, run_id)
        if not run:
            raise ValueError(f"Run {run_id} não encontrado")

        file_path = run.get("source_file_path")
        if not file_path or not Path(file_path).exists():
            raise ValueError("Arquivo original não encontrado. Reenvie o Excel.")

        log_event("EXCEL_IMPORT", "reanalyze_started", run_id=run_id, file_name=run["file_name"])
        staging.update_import_progress(
            db,
            run_id,
            status="ANALYZING",
            phase="loading_workbook",
            message="Relendo Excel original...",
        )

        _stage_workbook_rows(db, run_id, file_path)

        staging.update_import_progress(
            db,
            run_id,
            status="MAPPING_REQUIRED",
            phase="completed",
            message="Reanálise concluída. Revise o mapeamento e gere o preview.",
        )
        staging.update_run_status(db, run_id, "MAPPING_REQUIRED")
        db.commit()

        duration_ms = int((time.time() - started) * 1000)
        log_event("EXCEL_IMPORT", "reanalyze_completed", run_id=run_id, duration_ms=duration_ms)
    except Exception as exc:
        db.rollback()
        staging.update_import_progress(db, run_id, status="FAILED", phase="failed", message=str(exc))
        staging.update_run_status(db, run_id, "FAILED", error_message=str(exc))
        db.commit()
        log_error("EXCEL_IMPORT", "reanalyze_failed", run_id=run_id, error=str(exc), error_type=type(exc).__name__)
        raise
    finally:
        release_run_lock(run_id)
        db.close()


def start_rebuild_preview(db, run_id: int) -> dict[str, Any]:
    require_import_schema(db)

    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")
    if run["status"] == "APPLIED":
        raise ValueError("Importação já aplicada.")

    current_status = run["status"]
    log_event(
        "EXCEL_IMPORT",
        "preview_request_received",
        run_id=run_id,
        current_status=current_status,
        rebuild=True,
    )

    if current_status == "PREVIEWING":
        log_event(
            "EXCEL_IMPORT",
            "preview_request_ignored_already_running",
            run_id=run_id,
            current_status=current_status,
            rebuild=True,
        )
        return {
            "ok": True,
            "run_id": run_id,
            "status": "PREVIEWING",
            "already_running": True,
            "message": "Preview já está em processamento.",
        }

    if current_status in {"APPLYING", "APPLIED", "ANALYZING"}:
        raise ValueError(f"Processamento em andamento (status={current_status})")

    _ensure_staging_not_corrupted(db, run_id)

    if not try_acquire_run_lock(run_id):
        return {
            "ok": True,
            "run_id": run_id,
            "status": current_status,
            "already_running": True,
            "message": "Já existe processamento em andamento para esta importação.",
        }

    try:
        locked = db.execute(
            text(
                """
                UPDATE app_import_runs
                SET status = 'PREVIEWING',
                    phase = 'queued',
                    progress_message = 'Recalculando preview...'
                WHERE id = :run_id
                  AND status NOT IN ('PREVIEWING', 'APPLYING', 'APPLIED', 'ANALYZING')
                RETURNING id
                """
            ),
            {"run_id": run_id},
        ).scalar()
        if not locked:
            db.rollback()
            return {
                "ok": True,
                "run_id": run_id,
                "status": "PREVIEWING",
                "already_running": True,
                "message": "Preview já está em processamento.",
            }

        staging.update_import_progress(
            db,
            run_id,
            status="PREVIEWING",
            phase="queued",
            current=0,
            message="Recalculando preview...",
        )
        db.commit()
        log_event("EXCEL_IMPORT", "preview_job_starting", run_id=run_id, rebuild=True)
        submit_import_job("excel_preview", run_id, run_preview_job, run_id)
        return {
            "ok": True,
            "run_id": run_id,
            "status": "PREVIEWING",
            "message": "Preview será recalculado com o parser corrigido.",
        }
    except Exception:
        release_run_lock(run_id)
        raise


def start_apply(db, run_id: int, mode: str = "valid_rows_only") -> dict[str, Any]:
    validation = validate_preview_ready_for_apply(db, run_id, mode=mode)
    _ensure_staging_not_corrupted(db, run_id)

    if validation["run"]["status"] == "APPLYING":
        return {
            "ok": True,
            "run_id": run_id,
            "status": "APPLYING",
            "already_running": True,
            "message": "Apply já está em andamento.",
        }

    staging.update_import_progress(
        db,
        run_id,
        status="APPLYING",
        phase="queued",
        current=0,
        total=validation["valid_row_count"],
        message="Apply enfileirado...",
    )
    db.commit()
    submit_import_job("excel_apply", run_id, run_apply_job, run_id, mode)
    log_event("EXCEL_IMPORT", "apply_queued", run_id=run_id, mode=mode)
    return {
        "ok": True,
        "run_id": run_id,
        "status": "APPLYING",
        "mode": mode,
        "message": "Apply iniciado.",
    }


def run_apply_job(run_id: int, mode: str = "valid_rows_only") -> None:
    db = _job_db()
    started = time.time()
    try:
        run = staging.get_import_run(db, run_id)
        if not run:
            raise ValueError("Importação não encontrada.")

        validate_preview_ready_for_apply(db, run_id, mode=mode)

        log_event("EXCEL_IMPORT", "apply_started", run_id=run_id, mode=mode)

        mappings = staging.get_column_mappings(db, run_id)
        nullable_map = _column_nullable_map(db)
        column_meta_map = _column_metadata_map(db)
        total = staging.count_staging_rows(db, run_id)
        applied_rows = 0
        inserted = 0
        updated = 0
        unchanged = 0
        catalog_items_created = 0
        spec_links_created = 0
        spec_links_updated = 0
        skipped_error_rows = 0

        staging.update_import_progress(
            db,
            run_id,
            status="APPLYING",
            phase="applying",
            current=0,
            total=total,
            message="Aplicando alterações no banco...",
        )

        offset = 0
        batch_size = STAGING_BATCH
        while True:
            rows = staging.get_staging_rows(db, run_id, offset=offset, limit=batch_size)
            if not rows:
                break

            for row in rows:
                if row["row_status"] in {"ROW_ERROR", "ERROR", "SKIPPED_ERROR"}:
                    skipped_error_rows += 1
                    log_event(
                        "EXCEL_IMPORT",
                        "apply_row_skipped",
                        run_id=run_id,
                        excel_row_number=row["excel_row_number"],
                        row_status=row["row_status"],
                        reason=row.get("error_message") or "Erro crítico na linha.",
                    )
                    staging.record_apply_log(
                        db,
                        import_run_id=run_id,
                        action_type="SKIPPED_ERROR",
                        target_id=row.get("target_id"),
                        column_name=None,
                        old_value=str(row["excel_row_number"]),
                        new_value=row.get("error_message"),
                    )
                    continue
                if row["row_status"] == "UNCHANGED":
                    unchanged += 1
                    applied_rows += 1
                    continue

                raw_json = row["raw_json"]
                if isinstance(raw_json, str):
                    raw_json = json.loads(raw_json)

                target_id = row["target_id"]
                is_update = row["row_status"] == "UPDATE"

                mapped, row_errors, _, _, parse_metas = _build_mapped_row(
                    raw_json,
                    mappings,
                    nullable_map,
                    is_update=is_update,
                    run_id=run_id,
                    excel_row_number=row["excel_row_number"],
                    column_meta_map=column_meta_map,
                )
                if row_errors:
                    skipped_error_rows += 1
                    log_event(
                        "EXCEL_IMPORT",
                        "apply_row_skipped",
                        run_id=run_id,
                        excel_row_number=row["excel_row_number"],
                        reason="; ".join(item["error_message"] for item in row_errors),
                    )
                    staging.update_staging_row_preview(
                        db,
                        row_id=row["id"],
                        target_id=row.get("target_id"),
                        row_status="SKIPPED_ERROR",
                        error_message="; ".join(item["error_message"] for item in row_errors),
                    )
                    continue

                for col, val in mapped.items():
                    if col in NUMERIC_COLUMNS and val is not None:
                        try:
                            validate_numeric_range(val, col)
                        except ValueError as exc:
                            meta = parse_metas.get(col, {})
                            staging.record_import_error(
                                db,
                                import_run_id=run_id,
                                import_row_id=row["id"],
                                excel_row_number=row["excel_row_number"],
                                column_name=col,
                                value=meta.get("raw_value") or str(val),
                                error_message=str(exc),
                            )
                            raise ValueError(
                                f"Linha {row['excel_row_number']}, coluna {col}: "
                                f"valor {val} fora do limite. Verifique parser decimal."
                            ) from exc

                now = utc_now()

                try:
                    if row["row_status"] == "INSERT":
                        if target_id is None:
                            target_id = spec_items_service.get_next_id(db)
                        insert_data = {k: v for k, v in mapped.items() if k != "id"}
                        if "created_at" not in insert_data:
                            insert_data["created_at"] = now
                        if "updated_at" not in insert_data:
                            insert_data["updated_at"] = now
                        upsert_result = spec_items_service.upsert_spec_item(
                            db, int(target_id), insert_data, is_update=False
                        )
                        if upsert_result:
                            if upsert_result.get("catalog_item_created"):
                                catalog_items_created += 1
                            if upsert_result.get("spec_link_created"):
                                spec_links_created += 1
                        staging.record_apply_log(db, import_run_id=run_id, action_type="INSERT", target_id=int(target_id))
                        inserted += 1
                    elif row["row_status"] == "UPDATE" and target_id is not None:
                        update_data = {k: v for k, v in mapped.items() if k not in ("id", "created_at", "updated_at")}
                        update_data["updated_at"] = now
                        existing = spec_items_service.get_spec_item_by_id(db, int(target_id)) or {}
                        merged_payload = {**existing, **update_data}
                        upsert_result = spec_items_service.upsert_spec_item(
                            db, int(target_id), merged_payload, is_update=True
                        )
                        if upsert_result:
                            if upsert_result.get("catalog_item_created"):
                                catalog_items_created += 1
                            if upsert_result.get("spec_link_updated"):
                                spec_links_updated += 1
                        for col, new_val in update_data.items():
                            old_str = serialize_for_compare(existing.get(col))
                            new_str = serialize_for_compare(new_val)
                            if old_str != new_str:
                                staging.record_apply_log(
                                    db,
                                    import_run_id=run_id,
                                    action_type="UPDATE",
                                    target_id=int(target_id),
                                    column_name=col,
                                    old_value=old_str,
                                    new_value=new_str,
                                )
                        updated += 1
                except Exception as exc:
                    if "NumericValueOutOfRange" in type(exc).__name__ or "numeric" in str(exc).lower():
                        raise ValueError(
                            f"Linha {row['excel_row_number']}: valor numérico fora do limite "
                            f"numeric(18,6). Verifique parser decimal."
                        ) from exc
                    raise

                applied_rows += 1
                if applied_rows % PROGRESS_LOG_EVERY == 0:
                    log_event("EXCEL_IMPORT", "apply_progress", run_id=run_id, applied_rows=applied_rows)
                    staging.update_import_progress(
                        db,
                        run_id,
                        current=applied_rows,
                        total=total,
                        message=f"Aplicando... ({applied_rows}/{total})",
                        commit=True,
                    )

            offset += batch_size
            db.commit()

        staging.update_run_counts(
            db,
            run_id,
            inserted_rows=inserted,
            updated_rows=updated,
            unchanged_rows=unchanged,
            error_rows=run["error_rows"],
            status="APPLIED",
            applied=True,
        )
        staging.update_import_progress(
            db,
            run_id,
            status="APPLIED",
            phase="completed",
            current=total,
            total=total,
            message=(
                f"Importação aplicada. {applied_rows} linha(s) processada(s), "
                f"{skipped_error_rows} ignorada(s) por erro. "
                f"Canônicos criados: {catalog_items_created}, "
                f"vínculos criados: {spec_links_created}, "
                f"vínculos atualizados: {spec_links_updated}."
            ),
        )
        db.commit()

        duration_ms = int((time.time() - started) * 1000)
        log_event(
            "EXCEL_IMPORT",
            "apply_completed",
            run_id=run_id,
            applied_rows=applied_rows,
            skipped_error_rows=skipped_error_rows,
            inserted_rows=inserted,
            updated_rows=updated,
            unchanged_rows=unchanged,
            catalog_items_created=catalog_items_created,
            spec_links_created=spec_links_created,
            spec_links_updated=spec_links_updated,
            duration_ms=duration_ms,
        )
    except Exception as exc:
        db.rollback()
        staging.update_import_progress(db, run_id, status="FAILED", phase="failed", message=str(exc))
        staging.update_run_status(db, run_id, "FAILED", error_message=str(exc))
        db.commit()
        log_error("EXCEL_IMPORT", "apply_failed", run_id=run_id, error=str(exc), error_type=type(exc).__name__)
        raise
    finally:
        db.close()
