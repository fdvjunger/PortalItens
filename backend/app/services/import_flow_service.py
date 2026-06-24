import io
import json
import time
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.logging import log_error, log_event
from app.services import import_staging_service as staging
from app.services import spec_items_service
from app.services.import_guard import corrupted_staging_detail
from app.services.import_schema import require_import_schema
from app.services.import_staging_reader import build_raw_row_json, is_empty_excel_row
from app.utils.column_mapping_utils import normalize_header, suggest_column_mapping
from app.utils.excel_utils import (
    convert_value_for_import,
    detect_sheet_name,
    normalize_null,
    serialize_for_compare,
    utc_now,
)
from app.utils.value_parser import DATETIME_COLUMNS, NUMERIC_COLUMNS, try_parse_datetime_for_import
from app.core.db_schema import read_source_sql
from app.services.import_guard import validate_preview_ready_for_apply
from app.services.import_rules import critical_row_error_message, is_critical_row_error
from app.utils.db_coercion import NPS_ROW_CONTEXT_COLUMNS, PIPE_ROW_PEER_COLUMNS, build_coercion_warning_message


def _column_nullable_map(db: Session) -> dict[str, bool]:
    metadata = spec_items_service.get_column_metadata(db)
    return {col["column_name"]: col["is_nullable"] for col in metadata}


def _column_metadata_map(db: Session) -> dict[str, dict[str, Any]]:
    metadata = spec_items_service.get_column_metadata(db)
    return {col["column_name"]: col for col in metadata}


def _read_workbook(file_bytes: bytes) -> tuple[str, list[tuple[Any, ...]]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = detect_sheet_name(sheet_names)
    if not sheet_name:
        raise ValueError(
            f"Não foi possível detectar aba de dados. Abas encontradas: {', '.join(sheet_names)}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia.")
    return sheet_name, rows


def analyze_excel_file(db: Session, file_bytes: bytes, file_name: str) -> dict[str, Any]:
    started = time.time()
    log_event("IMPORT_EXCEL", "analyze started", file_name=file_name)

    try:
        sheet_name, rows = _read_workbook(file_bytes)
        raw_headers = [normalize_header(h) if h is not None else "" for h in rows[0]]
        db_columns = spec_items_service.get_valid_column_names(db)
        db_set = set(db_columns)

        excel_columns: list[str] = []
        for header in raw_headers:
            if header:
                excel_columns.append(header)

        if not excel_columns:
            raise ValueError("Nenhuma coluna encontrada no cabeçalho.")

        known_columns: list[str] = []
        unknown_columns: list[str] = []
        suggested_mappings: list[dict[str, Any]] = []
        mapping_records: list[dict[str, Any]] = []

        for excel_col in excel_columns:
            suggestion = suggest_column_mapping(excel_col, db_columns)
            if excel_col in db_set:
                known_columns.append(excel_col)
                mapping_records.append(
                    {
                        "excel_column_name": excel_col,
                        "target_column_name": excel_col,
                        "action": "MAP_TO_EXISTING",
                        "confidence": 1.0,
                    }
                )
            else:
                unknown_columns.append(excel_col)
                suggested_mappings.append(suggestion)
                mapping_records.append(
                    {
                        "excel_column_name": excel_col,
                        "target_column_name": suggestion.get("suggested_target_column"),
                        "action": suggestion.get("action", "IGNORE"),
                        "confidence": suggestion.get("confidence"),
                    }
                )

        data_rows = rows[1:]
        staging_rows: list[dict[str, Any]] = []
        for row_idx, row in enumerate(data_rows, start=2):
            if is_empty_excel_row(row):
                continue
            row_data = build_raw_row_json(raw_headers, row)
            staging_rows.append(
                {
                    "excel_row_number": row_idx,
                    "raw_json": row_data,
                }
            )

        run_id = staging.create_import_run(
            db,
            file_name=file_name,
            sheet_name=sheet_name,
            status="ANALYZED",
            total_rows=len(staging_rows),
        )
        staging.save_staging_rows(db, run_id, staging_rows)
        staging.save_column_mappings(db, run_id, mapping_records)
        db.commit()

        requires_mapping = any(
            m["action"] == "IGNORE" or m.get("confidence", 0) < 1.0
            for m in mapping_records
        )

        duration_ms = int((time.time() - started) * 1000)
        log_event(
            "IMPORT_EXCEL",
            "analyze completed",
            run_id=run_id,
            file_name=file_name,
            sheet_name=sheet_name,
            total_rows=len(staging_rows),
            known_columns_count=len(known_columns),
            unknown_columns_count=len(unknown_columns),
            duration_ms=duration_ms,
        )

        return {
            "ok": True,
            "run_id": run_id,
            "file_name": file_name,
            "sheet_name": sheet_name,
            "total_rows": len(staging_rows),
            "known_columns": known_columns,
            "unknown_columns": unknown_columns,
            "suggested_mappings": suggested_mappings,
            "requires_mapping": requires_mapping,
        }
    except Exception as exc:
        log_error("IMPORT_EXCEL", "analyze failed", file_name=file_name, error=str(exc), phase="analyze")
        raise


def get_import_run_detail(db: Session, run_id: int) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")

    staging_summary: dict[str, Any] = {}
    mapping_summary: dict[str, Any] = {}
    diff_count = 0
    error_count = 0

    try:
        row_counts = db.execute(
            text(
                """
                SELECT
                    COUNT(*) AS total_staged_rows,
                    COUNT(*) FILTER (WHERE row_status = 'PENDING') AS pending_rows,
                    COUNT(*) FILTER (WHERE row_status = 'INSERT') AS insert_rows,
                    COUNT(*) FILTER (WHERE row_status = 'UPDATE') AS update_rows,
                    COUNT(*) FILTER (WHERE row_status = 'UNCHANGED') AS unchanged_rows,
                    COUNT(*) FILTER (WHERE row_status = 'ERROR') AS staged_error_rows
                FROM app_import_rows
                WHERE import_run_id = :run_id
                """
            ),
            {"run_id": run_id},
        ).mappings().first()
        staging_summary = dict(row_counts or {})
    except Exception as exc:
        log_error(
            "EXCEL_IMPORT",
            "get_run_detail_staging_summary_failed",
            run_id=run_id,
            error=str(exc),
            error_type=type(exc).__name__,
        )

    try:
        mapping_counts = db.execute(
            text(
                """
                SELECT
                    COUNT(*) AS total_mappings,
                    COUNT(*) FILTER (WHERE action = 'MAP_TO_EXISTING') AS mapped_columns,
                    COUNT(*) FILTER (WHERE action = 'IGNORE') AS ignored_columns
                FROM app_import_column_mappings
                WHERE import_run_id = :run_id
                """
            ),
            {"run_id": run_id},
        ).mappings().first()
        mapping_summary = dict(mapping_counts or {})
    except Exception as exc:
        log_error(
            "EXCEL_IMPORT",
            "get_run_detail_mapping_summary_failed",
            run_id=run_id,
            error=str(exc),
            error_type=type(exc).__name__,
        )

    try:
        diff_count = int(
            db.execute(
                text("SELECT COUNT(*) FROM app_import_diffs WHERE import_run_id = :run_id"),
                {"run_id": run_id},
            ).scalar()
            or 0
        )
    except Exception as exc:
        log_error(
            "EXCEL_IMPORT",
            "get_run_detail_diff_count_failed",
            run_id=run_id,
            error=str(exc),
            error_type=type(exc).__name__,
        )

    try:
        error_count = int(
            db.execute(
                text("SELECT COUNT(*) FROM app_import_errors WHERE import_run_id = :run_id"),
                {"run_id": run_id},
            ).scalar()
            or 0
        )
    except Exception as exc:
        log_error(
            "EXCEL_IMPORT",
            "get_run_detail_error_count_failed",
            run_id=run_id,
            error=str(exc),
            error_type=type(exc).__name__,
        )

    return {
        "ok": True,
        "run": run,
        "staging_summary": staging_summary,
        "mapping_summary": mapping_summary,
        "diff_count": diff_count,
        "error_count": error_count,
    }


def get_run_with_mappings(db: Session, run_id: int) -> dict[str, Any]:
    detail = get_import_run_detail(db, run_id)
    mappings = staging.get_column_mappings(db, run_id)
    db_columns = spec_items_service.get_valid_column_names(db)

    mapping_rows = [
        {
            **mapping,
            "sample_values": [],
            "status": "mapped" if mapping["action"] == "MAP_TO_EXISTING" else "pending",
        }
        for mapping in mappings
    ]

    return {
        **detail,
        "column_mappings": mapping_rows,
        "target_columns": db_columns,
        "preview_summary": {
            "diff_count": detail.get("diff_count", 0),
            "error_count": detail.get("error_count", 0),
        },
    }


def save_mapping_and_preview(
    db: Session,
    run_id: int,
    mappings: list[dict[str, Any]],
) -> dict[str, Any]:
    require_import_schema(db)

    started = time.time()
    log_event("IMPORT_EXCEL", "mapping started", run_id=run_id, mappings_count=len(mappings))

    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")
    if run["status"] == "APPLIED":
        raise ValueError("Importação já aplicada.")

    db_columns = set(spec_items_service.get_valid_column_names(db))
    used_targets: set[str] = set()

    normalized_mappings: list[dict[str, Any]] = []
    for item in mappings:
        action = item.get("action", "IGNORE")
        excel_col = item["excel_column_name"]
        target = item.get("target_column_name")

        if action == "MAP_TO_EXISTING":
            if not target or target not in db_columns:
                raise ValueError(f"Coluna destino inválida para '{excel_col}'")
            if target in used_targets:
                raise ValueError(f"Coluna destino duplicada: {target}")
            used_targets.add(target)

        normalized_mappings.append(
            {
                "excel_column_name": excel_col,
                "target_column_name": target if action == "MAP_TO_EXISTING" else None,
                "action": action,
                "confidence": item.get("confidence"),
            }
        )

    staging.save_column_mappings(db, run_id, normalized_mappings)
    staging.clear_preview_artifacts(db, run_id)
    preview = _generate_preview(db, run_id)
    staging.update_run_counts(
        db,
        run_id,
        inserted_rows=preview["summary"]["insert_rows"],
        updated_rows=preview["summary"]["update_rows"],
        unchanged_rows=preview["summary"]["unchanged_rows"],
        ignored_rows=preview["summary"]["ignored_rows"],
        error_rows=preview["summary"]["error_rows"],
        status="PREVIEW_READY",
    )
    staging.save_preview_meta(db, run_id, preview.get("preview_meta", {}))
    db.commit()

    duration_ms = int((time.time() - started) * 1000)
    log_event(
        "IMPORT_EXCEL",
        "mapping completed",
        run_id=run_id,
        duration_ms=duration_ms,
        summary=preview["summary"],
    )
    return preview


def _format_row_error(excel_row_number: int | None, column_name: str, message: str) -> str:
    if excel_row_number is not None:
        return f"Linha {excel_row_number}, coluna {column_name}: {message}"
    return f"Coluna {column_name}: {message}"


def _diff_warning_message(column_name: str, col_meta: dict[str, Any] | None) -> str | None:
    if not col_meta:
        return None
    return build_coercion_warning_message(
        column_name,
        raw_value=col_meta.get("raw_value"),
        parsed_value=col_meta.get("parsed_value"),
        coercion_method=col_meta.get("coercion_method"),
        scale_divisor=col_meta.get("scale_divisor"),
        fallback_warning=col_meta.get("warning_message"),
    )


def _preview_response(
    db: Session,
    run_id: int,
    *,
    preview_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")

    mappings = staging.get_column_mappings(db, run_id)
    mapped_count = sum(1 for m in mappings if m["action"] == "MAP_TO_EXISTING")
    ignored_count = sum(1 for m in mappings if m["action"] == "IGNORE")

    meta = preview_meta or run.get("preview_meta") or {}
    if isinstance(meta, str):
        meta = json.loads(meta)

    valid_rows = (run["inserted_rows"] or 0) + (run["updated_rows"] or 0) + (run["unchanged_rows"] or 0)
    row_error_rows = run["error_rows"] or 0
    summary = _build_preview_summary(
        total_rows=run["total_rows"],
        insert_rows=run["inserted_rows"],
        update_rows=run["updated_rows"],
        unchanged_rows=run["unchanged_rows"],
        ignored_rows=run["ignored_rows"],
        row_error_rows=row_error_rows,
        warning_rows=staging.count_warning_rows(db, run_id),
        coerced_values=meta.get("coerced_values_count", len(meta.get("coerced_value_samples", []))),
        fatal_errors=meta.get("fatal_errors", 0),
    )

    return {
        "ok": True,
        "run_id": run_id,
        "summary": summary,
        "columns": {
            "mapped_count": mapped_count,
            "ignored_count": ignored_count,
            "unknown_count": 0,
        },
        "numeric_parse_samples": meta.get("numeric_parse_samples", {}),
        "coerced_value_samples": meta.get("coerced_value_samples", []),
        "sample_diffs": meta.get("sample_diffs") or staging.get_preview_diffs(db, run_id, limit=100),
        "warnings": staging.get_preview_warnings(db, run_id, limit=100),
        "errors": [],
        "warnings": [],
        "errors_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/errors",
        "warnings_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/warnings",
        "diffs_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/diffs",
    }


def get_preview(db: Session, run_id: int) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")

    status = run["status"]
    empty_urls = {
        "errors_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/errors",
        "warnings_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/warnings",
        "diffs_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/diffs",
    }

    if status == "CORRUPTED_STAGING":
        message = (
            run.get("error_message")
            or run.get("progress_message")
            or (
                "Foram detectados valores numéricos incompatíveis no staging. "
                "Use o diagnóstico para confirmar se o arquivo Excel já veio corrompido "
                "ou se houve erro durante o staging."
            )
        )
        return {
            "ok": False,
            "run_id": run_id,
            "status": "CORRUPTED_STAGING",
            "message": message,
            "hint": "Reenvie o arquivo ou use o botão Reanalisar arquivo.",
            "summary": None,
            "columns": {"mapped_count": 0, "ignored_count": 0, "unknown_count": 0},
            "numeric_parse_samples": {},
            "sample_diffs": [],
            "errors": [],
            "warnings": [],
            **empty_urls,
        }

    if status in {"PREVIEWING", "ANALYZING", "APPLYING"}:
        return {
            "ok": False,
            "run_id": run_id,
            "status": status,
            "message": "Preview ainda não está pronto.",
            "summary": None,
            "columns": {"mapped_count": 0, "ignored_count": 0, "unknown_count": 0},
            "numeric_parse_samples": {},
            "sample_diffs": [],
            "errors": [],
            "warnings": [],
            **empty_urls,
        }

    if status == "MAPPING_REQUIRED":
        return {
            "ok": False,
            "run_id": run_id,
            "status": "MAPPING_REQUIRED",
            "message": "Preview ainda não foi gerado.",
            "summary": None,
            "columns": {"mapped_count": 0, "ignored_count": 0, "unknown_count": 0},
            "numeric_parse_samples": {},
            "sample_diffs": [],
            "errors": [],
            "warnings": [],
            **empty_urls,
        }

    if status not in {"PREVIEW_READY", "APPLIED", "FAILED"}:
        return {
            "ok": False,
            "run_id": run_id,
            "status": status,
            "message": "Preview indisponível neste status.",
            "summary": None,
            "columns": {"mapped_count": 0, "ignored_count": 0, "unknown_count": 0},
            "numeric_parse_samples": {},
            "sample_diffs": [],
            "errors": [],
            "warnings": [],
            **empty_urls,
        }

    return _preview_response(db, run_id)


def _build_nps_row_context(
    raw_json: dict[str, Any],
    mappings: list[dict[str, Any]],
) -> dict[str, Any]:
    context: dict[str, Any] = {}
    for mapping in mappings:
        if mapping["action"] != "MAP_TO_EXISTING":
            continue
        target_col = mapping["target_column_name"]
        if target_col in NPS_ROW_CONTEXT_COLUMNS:
            context[target_col] = raw_json.get(mapping["excel_column_name"])
    return context


def _build_mapped_row(
    raw_json: dict[str, Any],
    mappings: list[dict[str, Any]],
    nullable_map: dict[str, bool],
    *,
    is_update: bool,
    run_id: int | None = None,
    excel_row_number: int | None = None,
    stats: dict[str, int] | None = None,
    numeric_parse_samples: dict[str, list[dict[str, Any]]] | None = None,
    column_meta_map: dict[str, dict[str, Any]] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[str], list[tuple[str, str, str | None, str | None]], dict[str, dict[str, Any]]]:
    mapped: dict[str, Any] = {}
    row_errors: list[dict[str, Any]] = []
    warnings: list[str] = []
    diffs: list[tuple[str, str, str | None, str | None]] = []
    parse_metas: dict[str, dict[str, Any]] = {}

    row_peers: dict[str, Any] = {}
    for mapping in mappings:
        if mapping["action"] != "MAP_TO_EXISTING":
            continue
        target_col = mapping["target_column_name"]
        if target_col in PIPE_ROW_PEER_COLUMNS:
            row_peers[target_col] = raw_json.get(mapping["excel_column_name"])

    nps_row_context = _build_nps_row_context(raw_json, mappings)

    for mapping in mappings:
        if mapping["action"] != "MAP_TO_EXISTING":
            continue

        excel_col = mapping["excel_column_name"]
        target_col = mapping["target_column_name"]
        if not target_col:
            continue

        raw_value = raw_json.get(excel_col)

        if is_update and target_col in DATETIME_COLUMNS:
            if target_col == "created_at" and raw_value is not None and str(raw_value).strip():
                warnings.append("created_at não será atualizado (valor Excel ignorado)")
            if target_col == "updated_at" and raw_value is not None and str(raw_value).strip():
                warnings.append("updated_at será definido automaticamente com now() no apply")
            if stats is not None:
                stats["datetimes_skipped_update"] = stats.get("datetimes_skipped_update", 0) + 1
            continue

        if target_col in DATETIME_COLUMNS:
            converted, error, warning, status = try_parse_datetime_for_import(
                target_col, raw_value, is_update=False
            )
            if stats is not None:
                if status == "parsed":
                    stats["datetimes_normalized"] = stats.get("datetimes_normalized", 0) + 1
                elif status in {"ignored", "warning"}:
                    stats["datetimes_ignored"] = stats.get("datetimes_ignored", 0) + 1
            if warning:
                warnings.append(warning)
            if converted is not None and converted != "__SKIP__":
                mapped[target_col] = converted
            continue

        converted, error, warning, meta = convert_value_for_import(
            target_col,
            raw_value,
            is_update=is_update,
            is_nullable=nullable_map.get(target_col, True),
            run_id=run_id,
            excel_row_number=excel_row_number,
            column_meta=(column_meta_map or {}).get(target_col),
            row_peers=row_peers,
            row_context=nps_row_context,
        )

        if meta:
            parse_metas[target_col] = meta
            if stats is not None and meta.get("coercion_method") in {
                "NUMERIC_SCALE_INFERRED",
                "NULL_NUMERIC_OUT_OF_PROFILE",
                "NULL_PARSE_FAILED",
            }:
                stats["coerced_values"] = stats.get("coerced_values", 0) + 1
            if numeric_parse_samples is not None and target_col in NUMERIC_COLUMNS:
                samples = numeric_parse_samples.setdefault(target_col, [])
                if len(samples) < 3 and meta.get("raw_value") is not None:
                    sample = {
                        "excel_row_number": excel_row_number,
                        "raw": meta.get("raw_value"),
                        "parsed": meta.get("parsed_value"),
                        "coercion_method": meta.get("coercion_method"),
                        "warning_message": meta.get("warning_message"),
                    }
                    samples.append(sample)
                    log_event(
                        "EXCEL_IMPORT",
                        "numeric_parse_sample",
                        run_id=run_id,
                        excel_row_number=excel_row_number,
                        column_name=target_col,
                        raw_value=meta.get("raw_value"),
                        parsed_value=meta.get("parsed_value"),
                    )

        if error:
            if is_critical_row_error(target_col, raw_value, error):
                row_errors.append(
                    {
                        "column_name": target_col,
                        "raw_value": (meta or {}).get("raw_value") or str(raw_value),
                        "error_message": critical_row_error_message(target_col, error),
                    }
                )
            else:
                warnings.append(error)
            continue
        if warning:
            warnings.append(warning)
        if converted == "__SKIP__":
            continue

        if converted is not None or target_col in mapped:
            mapped[target_col] = converted

    return mapped, row_errors, warnings, diffs, parse_metas


def _build_preview_summary(
    *,
    total_rows: int,
    insert_rows: int,
    update_rows: int,
    unchanged_rows: int,
    ignored_rows: int,
    row_error_rows: int,
    warning_rows: int,
    coerced_values: int = 0,
    fatal_errors: int = 0,
) -> dict[str, Any]:
    valid_rows = insert_rows + update_rows + unchanged_rows
    return {
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "row_error_rows": row_error_rows,
        "insert_rows": insert_rows,
        "update_rows": update_rows,
        "unchanged_rows": unchanged_rows,
        "ignored_rows": ignored_rows,
        "error_rows": row_error_rows,
        "warning_rows": warning_rows,
        "coerced_values": coerced_values,
        "fatal_errors": fatal_errors,
        "can_apply_valid_rows": valid_rows > 0 and fatal_errors == 0,
    }


def _generate_preview(
    db: Session,
    run_id: int,
    *,
    progress_callback=None,
) -> dict[str, Any]:
    require_import_schema(db)

    started = time.time()
    log_event("EXCEL_IMPORT", "preview_started", run_id=run_id)
    log_event(
        "EXCEL_IMPORT",
        "numeric_parser_rules",
        run_id=run_id,
        rules="intelligent_db_coercion;scale_inference;nullable_to_null",
    )

    staging.clear_preview_artifacts(db, run_id)
    db.commit()

    mappings = staging.get_column_mappings(db, run_id)
    nullable_map = _column_nullable_map(db)
    column_meta_map = _column_metadata_map(db)
    total_rows = staging.count_staging_rows(db, run_id)

    if progress_callback:
        progress_callback(0, total_rows, "Preparando geração do preview...")

    insert_rows = 0
    update_rows = 0
    unchanged_rows = 0
    ignored_rows = 0
    error_rows = 0
    row_error_rows = 0
    warning_rows = 0
    datetime_stats: dict[str, int] = {
        "datetimes_normalized": 0,
        "datetimes_ignored": 0,
        "datetimes_skipped_update": 0,
        "coerced_values": 0,
    }
    numeric_parse_samples: dict[str, list[dict[str, Any]]] = {}
    coerced_value_samples: list[dict[str, Any]] = []
    preview_sample_diffs: list[dict[str, Any]] = []

    excel_ids: list[int] = []
    id_mapping = next((m for m in mappings if m.get("target_column_name") == "id"), None)

    offset = 0
    batch_size = 500
    processed = 0
    parsed_meta: list[dict[str, Any]] = []

    while True:
        batch = staging.get_staging_rows(db, run_id, offset=offset, limit=batch_size)
        if not batch:
            break
        for row in batch:
            raw_json = row["raw_json"]
            if isinstance(raw_json, str):
                raw_json = json.loads(raw_json)

            excel_row_number = row["excel_row_number"]
            row_id: int | None = None
            errors: list[str] = []
            if id_mapping:
                raw_id = raw_json.get(id_mapping["excel_column_name"])
                converted, error, _, _ = convert_value_for_import(
                    "id",
                    raw_id,
                    is_update=False,
                    is_nullable=False,
                    run_id=run_id,
                    excel_row_number=excel_row_number,
                    column_meta=column_meta_map.get("id"),
                )
                if error:
                    errors.append(_format_row_error(excel_row_number, "id", error))
                elif converted is not None:
                    row_id = int(converted)

            parsed_meta.append({**row, "row_id": row_id, "errors": errors})
            if row_id is not None:
                excel_ids.append(row_id)

        offset += batch_size

    existing_ids = spec_items_service.get_existing_ids(db, excel_ids)
    max_id = db.execute(text(f"SELECT COALESCE(MAX(id), 0) FROM {read_source_sql()}")).scalar()
    next_auto_id = int(max_id) + 1

    for parsed in parsed_meta:
        processed += 1
        if processed == 1 or processed % 1000 == 0 or processed == total_rows:
            log_event("EXCEL_IMPORT", "preview_progress", run_id=run_id, processed_rows=processed)
            if processed % 1000 == 0:
                log_event("EXCEL_IMPORT", "numeric_parse_progress", run_id=run_id, processed_rows=processed)
            if progress_callback:
                progress_callback(processed, total_rows, f"Gerando preview... ({processed}/{total_rows})")

        row_id_db = parsed["id"]
        excel_row_number = parsed["excel_row_number"]
        raw_json = parsed["raw_json"]
        if isinstance(raw_json, str):
            raw_json = json.loads(raw_json)

        if parsed.get("errors"):
            row_error_rows += 1
            message = "; ".join(parsed["errors"])
            staging.update_staging_row_preview(
                db, row_id=row_id_db, target_id=None, row_status="ROW_ERROR", error_message=message
            )
            staging.record_import_error(
                db,
                import_run_id=run_id,
                import_row_id=row_id_db,
                excel_row_number=excel_row_number,
                column_name="id",
                error_message=message,
            )
            continue

        row_id = parsed.get("row_id")
        is_update = row_id is not None and row_id in existing_ids
        existing_item = (
            spec_items_service.get_spec_item_by_id(db, row_id) if is_update and row_id is not None else None
        )

        mapped, row_errors, warnings, _, parse_metas = _build_mapped_row(
            raw_json,
            mappings,
            nullable_map,
            is_update=is_update,
            run_id=run_id,
            excel_row_number=excel_row_number,
            stats=datetime_stats,
            numeric_parse_samples=numeric_parse_samples,
            column_meta_map=column_meta_map,
        )

        for _col, meta in parse_metas.items():
            method = meta.get("coercion_method")
            if (
                method
                and method not in {
                    "NUMERIC_DIRECT",
                    "TEXT_STRING",
                    "INTEGER_PARSED",
                    "BOOLEAN_PARSED",
                    "DATETIME_PARSED",
                    "NULL_EMPTY",
                }
                and len(coerced_value_samples) < 100
            ):
                coerced_value_samples.append(
                    {
                        "excel_row_number": excel_row_number,
                        "column_name": _col,
                        "raw_value": meta.get("raw_value"),
                        "parsed_value": meta.get("parsed_value"),
                        "coercion_method": meta.get("coercion_method"),
                        "scale_divisor": str(meta.get("scale_divisor")) if meta.get("scale_divisor") else None,
                        "warning_message": _diff_warning_message(_col, meta),
                    }
                )

        if warnings:
            warning_rows += 1

        if row_errors:
            row_error_rows += 1
            message = "; ".join(item["error_message"] for item in row_errors)
            staging.update_staging_row_preview(
                db, row_id=row_id_db, target_id=row_id, row_status="ROW_ERROR", error_message=message
            )
            for item in row_errors:
                staging.record_import_error(
                    db,
                    import_run_id=run_id,
                    import_row_id=row_id_db,
                    excel_row_number=excel_row_number,
                    column_name=item.get("column_name"),
                    value=item.get("raw_value"),
                    error_message=item.get("error_message"),
                )
            continue

        target_id = row_id
        row_status = "INSERT"
        saved_warning_messages: set[str] = set()

        if row_id is None:
            target_id = next_auto_id
            next_auto_id += 1
            insert_rows += 1
        elif is_update:
            has_changes = False
            for col, new_val in mapped.items():
                if col == "id":
                    continue
                old_val = existing_item.get(col) if existing_item else None
                old_str = serialize_for_compare(old_val)
                new_str = serialize_for_compare(new_val)
                if old_str != new_str:
                    has_changes = True
                    col_meta = parse_metas.get(col, {})
                    warn_msg = _diff_warning_message(col, col_meta)
                    if warn_msg:
                        saved_warning_messages.add(warn_msg)
                    staging.save_diff(
                        db,
                        run_id=run_id,
                        import_row_id=row_id_db,
                        excel_row_number=excel_row_number,
                        target_id=row_id,
                        column_name=col,
                        old_value=old_str,
                        new_value=new_str,
                        diff_type="UPDATE",
                        raw_value=col_meta.get("raw_value"),
                        parsed_value=col_meta.get("parsed_value"),
                        coercion_method=col_meta.get("coercion_method"),
                        scale_divisor=col_meta.get("scale_divisor"),
                        warning_message=warn_msg,
                    )
                    if len(preview_sample_diffs) < 100:
                        col_meta = parse_metas.get(col, {})
                        preview_sample_diffs.append(
                            {
                                "excel_row_number": excel_row_number,
                                "target_id": row_id,
                                "column_name": col,
                                "raw_value": col_meta.get("raw_value"),
                                "parsed_value": col_meta.get("parsed_value"),
                                "old_value": old_str,
                                "new_value": new_str,
                                "diff_type": "UPDATE",
                            }
                        )
            if has_changes:
                update_rows += 1
                row_status = "UPDATE"
            else:
                unchanged_rows += 1
                row_status = "UNCHANGED"
        else:
            insert_rows += 1
            for col, new_val in mapped.items():
                if col == "id":
                    continue
                new_str = serialize_for_compare(new_val)
                col_meta = parse_metas.get(col, {})
                warn_msg = _diff_warning_message(col, col_meta)
                if warn_msg:
                    saved_warning_messages.add(warn_msg)
                staging.save_diff(
                    db,
                    run_id=run_id,
                    import_row_id=row_id_db,
                    excel_row_number=excel_row_number,
                    target_id=target_id,
                    column_name=col,
                    old_value=None,
                    new_value=new_str,
                    diff_type="INSERT",
                    raw_value=col_meta.get("raw_value"),
                    parsed_value=col_meta.get("parsed_value"),
                    coercion_method=col_meta.get("coercion_method"),
                    scale_divisor=col_meta.get("scale_divisor"),
                    warning_message=warn_msg,
                )
                if len(preview_sample_diffs) < 100:
                    col_meta = parse_metas.get(col, {})
                    preview_sample_diffs.append(
                        {
                            "excel_row_number": excel_row_number,
                            "target_id": target_id,
                            "column_name": col,
                            "raw_value": col_meta.get("raw_value"),
                            "parsed_value": col_meta.get("parsed_value"),
                            "old_value": None,
                            "new_value": new_str,
                            "diff_type": "INSERT",
                        }
                    )

        for warning in warnings:
            if warning in saved_warning_messages:
                continue
            staging.save_diff(
                db,
                run_id=run_id,
                import_row_id=row_id_db,
                excel_row_number=excel_row_number,
                target_id=target_id,
                column_name="",
                old_value=None,
                new_value=None,
                diff_type="WARNING",
                warning_message=warning,
            )
            saved_warning_messages.add(warning)

        staging.update_staging_row_preview(
            db, row_id=row_id_db, target_id=target_id, row_status=row_status
        )

        if processed % 500 == 0:
            db.commit()

    db.commit()

    mapped_count = sum(1 for m in mappings if m["action"] == "MAP_TO_EXISTING")
    ignored_count = sum(1 for m in mappings if m["action"] == "IGNORE")
    duration_ms = int((time.time() - started) * 1000)

    log_event(
        "EXCEL_IMPORT",
        "preview_completed",
        run_id=run_id,
        insert_rows=insert_rows,
        update_rows=update_rows,
        unchanged_rows=unchanged_rows,
        error_rows=row_error_rows,
        warning_rows=warning_rows,
        datetimes_normalized=datetime_stats["datetimes_normalized"],
        datetimes_ignored=datetime_stats["datetimes_ignored"],
        duration_ms=duration_ms,
    )

    preview_meta = {
        "numeric_parse_samples": numeric_parse_samples,
        "sample_diffs": preview_sample_diffs,
        "coerced_value_samples": coerced_value_samples,
        "coerced_values_count": datetime_stats.get("coerced_values", 0),
        "fatal_errors": 0,
        "row_error_rows": row_error_rows,
        "valid_rows": insert_rows + update_rows + unchanged_rows,
    }
    staging.save_preview_meta(db, run_id, preview_meta)

    summary = _build_preview_summary(
        total_rows=total_rows,
        insert_rows=insert_rows,
        update_rows=update_rows,
        unchanged_rows=unchanged_rows,
        ignored_rows=ignored_rows,
        row_error_rows=row_error_rows,
        warning_rows=warning_rows,
        coerced_values=datetime_stats.get("coerced_values", 0),
        fatal_errors=0,
    )

    return {
        "ok": True,
        "run_id": run_id,
        "summary": summary,
        "columns": {
            "mapped_count": mapped_count,
            "ignored_count": ignored_count,
            "unknown_count": 0,
        },
        "numeric_parse_samples": numeric_parse_samples,
        "coerced_value_samples": coerced_value_samples,
        "sample_diffs": preview_sample_diffs,
        "preview_meta": preview_meta,
        "errors": [],
        "warnings": staging.get_preview_warnings(db, run_id, limit=100),
        "errors_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/errors",
        "warnings_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/warnings",
        "diffs_paginated_url": f"/api/spec-items/import-excel/runs/{run_id}/diffs",
    }


def apply_import(db: Session, run_id: int) -> dict[str, Any]:
    started = time.time()
    log_event("IMPORT_EXCEL", "apply started", run_id=run_id)

    validate_preview_ready_for_apply(db, run_id, mode="valid_rows_only")

    mappings = staging.get_column_mappings(db, run_id)
    nullable_map = _column_nullable_map(db)
    column_meta_map = _column_metadata_map(db)
    staging_rows = staging.get_staging_rows(db, run_id)

    inserted = 0
    updated = 0
    unchanged = 0

    try:
        for row in staging_rows:
            if row["row_status"] in {"ROW_ERROR", "ERROR", "SKIPPED_ERROR"}:
                continue
            if row["row_status"] == "UNCHANGED":
                unchanged += 1
                continue

            raw_json = row["raw_json"]
            if isinstance(raw_json, str):
                raw_json = json.loads(raw_json)

            target_id = row["target_id"]
            is_update = row["row_status"] == "UPDATE"

            mapped, row_errors, _, _, _parse_metas = _build_mapped_row(
                raw_json,
                mappings,
                nullable_map,
                is_update=is_update,
                run_id=run_id,
                excel_row_number=row["excel_row_number"],
                column_meta_map=column_meta_map,
            )
            if row_errors:
                continue

            now = utc_now()

            if row["row_status"] == "INSERT":
                if target_id is None:
                    target_id = spec_items_service.get_next_id(db)
                insert_data = {k: v for k, v in mapped.items() if k != "id"}
                if "created_at" not in insert_data:
                    insert_data["created_at"] = now
                if "updated_at" not in insert_data:
                    insert_data["updated_at"] = now
                upsert_result = spec_items_service.upsert_spec_item(
                    db, int(target_id), insert_data, is_update=False
                )
                staging.record_apply_log(
                    db,
                    import_run_id=run_id,
                    action_type="INSERT",
                    target_id=int(target_id),
                )
                inserted += 1
            elif row["row_status"] == "UPDATE" and target_id is not None:
                update_data = {
                    k: v
                    for k, v in mapped.items()
                    if k not in ("id", "created_at", "updated_at")
                }
                update_data["updated_at"] = now
                existing = spec_items_service.get_spec_item_by_id(db, int(target_id)) or {}
                merged_payload = {**existing, **update_data}
                spec_items_service.upsert_spec_item(
                    db, int(target_id), merged_payload, is_update=True
                )
                for col, new_val in update_data.items():
                    old_str = serialize_for_compare(existing.get(col))
                    new_str = serialize_for_compare(new_val)
                    if old_str != new_str:
                        staging.record_apply_log(
                            db,
                            import_run_id=run_id,
                            action_type="UPDATE",
                            target_id=int(target_id),
                            column_name=col,
                            old_value=old_str,
                            new_value=new_str,
                        )
                updated += 1

        staging.update_run_counts(
            db,
            run_id,
            inserted_rows=inserted,
            updated_rows=updated,
            unchanged_rows=unchanged,
            error_rows=run["error_rows"],
            status="APPLIED",
            applied=True,
        )
        db.commit()

        duration_ms = int((time.time() - started) * 1000)
        log_event(
            "IMPORT_EXCEL",
            "apply completed",
            run_id=run_id,
            inserted_rows=inserted,
            updated_rows=updated,
            unchanged_rows=unchanged,
            duration_ms=duration_ms,
        )

        return {
            "ok": True,
            "run_id": run_id,
            "status": "APPLIED",
            "inserted_rows": inserted,
            "updated_rows": updated,
            "unchanged_rows": unchanged,
            "error_rows": run["error_rows"],
        }
    except Exception as exc:
        db.rollback()
        staging.update_run_status(db, run_id, "FAILED", error_message=str(exc))
        staging.update_run_counts(db, run_id, status="FAILED")
        db.commit()
        log_error("IMPORT_EXCEL", "apply failed", run_id=run_id, error=str(exc), phase="apply")
        raise


def cancel_import(db: Session, run_id: int) -> dict[str, Any]:
    run = staging.get_import_run(db, run_id)
    if not run:
        raise ValueError("Importação não encontrada.")
    if run["status"] == "APPLIED":
        raise ValueError("Importação já aplicada não pode ser cancelada.")

    staging.update_run_status(db, run_id, "CANCELLED")
    db.commit()
    log_event("IMPORT_EXCEL", "import cancelled", run_id=run_id)
    return {"ok": True, "run_id": run_id, "status": "CANCELLED"}
