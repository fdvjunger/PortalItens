import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.services import spec_items_service


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _write_data_sheet(ws: Worksheet, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([_serialize_value(row.get(col)) for col in columns])

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for idx, _ in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18


def build_export_workbook(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "spec_items"
    _write_data_sheet(ws, columns, rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_template_workbook(db: Session) -> bytes:
    metadata = spec_items_service.get_column_metadata(db)
    columns = [col["column_name"] for col in metadata]

    wb = Workbook()
    ws_items = wb.active
    ws_items.title = "spec_items"
    ws_items.append(columns)
    for cell in ws_items[1]:
        cell.font = Font(bold=True)
    ws_items.freeze_panes = "A2"

    ws_meta = wb.create_sheet("metadata")
    ws_meta.append(["column_name", "data_type", "required", "notes"])
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)

    for col in metadata:
        required = "Não" if col["is_nullable"] else "Sim"
        notes = ""
        if col["column_name"] == "cliente":
            notes = "Cliente do item. Pode ser NULL. Sem valor default."
        if col["column_name"] == "id":
            notes = "Identificador único. Deixe vazio para gerar automaticamente na importação."
        ws_meta.append([col["column_name"], col["data_type"], required, notes])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
